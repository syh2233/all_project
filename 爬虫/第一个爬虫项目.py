import requests
from bs4 import BeautifulSoup
import time
import re
import json
import openpyxl

def read_url_from_excel(file_path, sheet_name):
    # 读取Excel文件中的链接
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    i = 0
    for i in range(5):
        url = sheet.cell(row=i + 1, column=1).value
    return url

class Spider(object):

    def __init__(self):
        # 从Excel文件中读取链接
        file_path = 'C:/Users/沈家/Desktop/cheshi.xlsx'
        sheet_name = 'Sheet1'
        self.url1 = read_url_from_excel(file_path, sheet_name)

    def get_content_from_url(self, url1):
        # 根据url，获取响应内容的字符串
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
        }
        content1 = requests.get(url1, headers=headers, timeout=2)
        content = content1.text
        print(content1.status_code)
        return content1.content.decode()

    def parse_2070(self, content):
        # 提取所要数据
        soup = BeautifulSoup(content, 'lxml')
        Sku = soup.findAll('script', {'type': 'text/javascript'})
        # print(Sku[2].text)
        sku = re.findall('selectedSKU": "(.+)",\\n  "cart_order_merged_profile":', Sku[2].text, re.DOTALL)[0]
        print(sku)
        img1 = soup.find(attrs={'class': 'thumb selected slick-slide'})
        img1_str = str(img1)
        img = re.findall(r'<a class="thumbnail-link" href="(.+)" tabindex="-1" target="_blank"', img1_str, re.DOTALL)[0]
        print(img)
        title = re.findall(r'<title>(.+)\| Talbots</title>', content, re.DOTALL)[0]
        print(title)
        price1 = soup.find(attrs={'data-currencysymbol': '$'})
        price = price1.text
        price = price.strip()
        print(price)
        color1 = soup.find(attrs={'class': 'selected-value'})
        color = color1.text
        color = color.strip()
        print(color)
        desc = soup.find(attrs={'class': 'product-description-content pdp-description'})
        print(desc.text)
        vule = [sku, title, price, color, desc.text]
        # vule_str = json.dumps(vule)
        # print(vule_str)
        return vule

    def img(self, content):
        soup = BeautifulSoup(content, 'lxml')
        img1 = soup.find(attrs={'class': 'thumb selected slick-slide'})
        img1_str = str(img1)
        img = re.findall(r'<a class="thumbnail-link" href="(.+)" tabindex="-1" target="_blank"', img1_str, re.DOTALL)[0]
        return img

    def save(self, vule, path):
        # 保存数据
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Sheet5']
        max_row = sheet.max_row
        print(max_row)
        with open(path, 'a', encoding='utf-8') as ft:
            ft.write(str(vule))  # 将列表转换为字符串
        for i in range(len(vule)):
            sheet.cell(max_row + 1, i + 1, vule[i])

    def save_data_to_excel(self, data, file_path):
        # 创建一个新的工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 将数据写入工作表
        for row in data:
            sheet.append(row)

        # 保存工作簿
        workbook.save(file_path)

    def download_image(self, img, save_path, ):
        # 保存图片
        response = requests.get(img)
        with open(save_path, 'wb') as f:
            f.write(response.content)

    def crawl_2070(self):
        # 运行程序
        self.__init__()
        content = self.get_content_from_url(self.url1)
        shuju = self.parse_2070(content)
        img = self.img(content)
        # 将shuju转换为列表
        shuju_list = list(shuju)
        self.save_data_to_excel(shuju_list, 'C:/Users/沈家/Desktop/cheshi.xlsx')
        self.download_image(img, 'E:/vulu[0].jpg')

    def run(self):
        self.crawl_2070()


if __name__ == '__main__':
    spider = Spider()
    spider.run()
