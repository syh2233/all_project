# import openpyxl
#
# workbook = openpyxl.load_workbook('C:/Users/沈家/Desktop/cheshi.xlsx')
# sheet = workbook['Sheet2']
# print(sheet)

# from openpyxl import  load_workbook
# import xlwt
# import openpyxl as px
#
# wb = load_workbook(r'C:\Users\沈家\Desktop\Content.xlsx')
#
# # sheet = wb.worksheets[0]
# sheet = wb['Content']
#
# """
# 行
# min_row
# max_row
# 列
# min_col
# max_col
# """
# rows = []
# for row in sheet.columns:
#     rows.append([cell.value for cell in row])
# print(rows)
#
# wb = px.Workbook()
# sheet1 = wb.active
# sheet1.title = "sheet5"
# for r, row in enumerate(rows):
#     for c, value in enumerate(row):
#         sheet1.cell(row=r+1, column=c+1, value=value)
# wb.save("2070.xlsx")
# rode = input("请输入路径:")
# rode_re = "E:\\BaiduSyncdisk\\2000\\" + rode + "\\" + rode + ".xlsx"
# print(rode_re)
# a = ord('L')-64
# print(a)

# #
# df = pd.read_excel(io=r'E:\BaiduSyncdisk\2000\2081\2081.xlsx')
# df1 = df['Category']
# print(df1)
# df1 = df['size'].str.split('|')
# dictionary = df1.to_dict()
# capitalized = {key: str(value).capitalize() for key, value in dictionary.items()}
# if len(capitalized[0])==len(capitalized[1]):
#     print(len(capitalized[0]))
# else:
#     print('no')
# lit = df1
# print(df1)
# for i, row1 in enumerate(range(1, 101, 10)):
#     row = row1 + 1
#     print(i, row)
#     # cell = worksheet.cell(row=row1 + 1, column=4)
    # cell.value = ('=VLOOKUP(AD' + str(i + 1) + ',Sheet4!A:B,2,FALSE)')

import requests
from bs4 import BeautifulSoup
import time
import re
import json
import openpyxl

class Spider(object):

    def __init__(self):
        self.url1 = "https://www.talbots.com/airy-gauze-button-front-shirt/P242075439.html?cgid=apparel-blouses-and-shirts&dwvar_P242075439_color=WHITE&dwvar_P242075439_sizeType=MS"

    def get_content_from_url(self, url1):
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
        }
        content1 = requests.get(url1, headers=headers, timeout=2)
        content = content1.text
        print(content1.status_code)
        return content1.content.decode()

    def parse_2070(self, content):
        soup = BeautifulSoup(content, 'lxml')
        Sku = soup.findAll('script', {'type': 'text/javascript'})
        print(Sku[2].text)
        sku = re.findall('selectedSKU": "(.+)",\\n  "cart_order_merged_profile":', Sku[2].text, re.DOTALL)[0]
        print(sku)
        img1 = soup.find(attrs={'class': 'thumb selected slick-slide'})
        img1_str = str(img1)
        img = re.findall(r'<a class="thumbnail-link" href="(.+)" tabindex="-1" target="_blank"', img1_str, re.DOTALL)[0]
        print(img)
        title = re.findall(r'<title>(.+)\| Talbots</title>', content, re.DOTALL)[0]
        print(title)
        price1 = soup.find(attrs={'data-currencysymbol': '$'})
        price = price1.text
        price = price.strip()
        print(price)
        color1 = soup.find(attrs={'class': 'selected-value'})
        color = color1.text
        color = color.strip()
        print(color)
        desc = soup.find(attrs={'class': 'product-description-content pdp-description'})
        print(desc.text)
        vule = [sku, title, price, color, desc.text]
        return vule

    def img(self, content):
        soup = BeautifulSoup(content, 'lxml')
        img1 = soup.find(attrs={'class': 'thumb selected slick-slide'})
        img1_str = str(img1)
        img = re.findall(r'<a class="thumbnail-link" href="(.+)" tabindex="-1" target="_blank"', img1_str, re.DOTALL)[0]
        return img

    def save(self, vule, path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook['Sheet5']
        max_row = sheet.max_row
        print(max_row)
        for i in range(len(vule)):
            sheet.cell(max_row + 1, i + 1, vule[i])
        workbook.save(path)

    def download_image(self, img, save_path):
        response = requests.get(img)
        with open(save_path, 'wb') as f:
            f.write(response.content)

    def crawl_2070(self, url):
        content = self.get_content_from_url(url)
        shuju = self.parse_2070(content)
        img = self.img(content)
        self.save(shuju, 'C:/Users/沈家/Desktop/cheshi.xlsx')
        self.download_image(img, 'E:/vulu[0].jpg')

    def run(self):
        workbook = openpyxl.load_workbook('C:/Users/沈家/Desktop/cheshi.xlsx')
        sheet = workbook['Sheet1']
        for row in range(2, sheet.max_row + 1):
            url = sheet.cell(row, 1).value
            self.crawl_2070(url)

if __name__ == '__main__':
    spider = Spider()
    spider.run()
